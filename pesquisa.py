import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import subprocess

file_name = 'cadastro_entregas.xlsx'

# Define os tipos de dados das colunas para garantir que todos os campos numéricos longos sejam tratados como strings
col_types = {
    'Produtor/Empresa': str,
    'CPF/CNPJ': str,
    'Fone': str,
    'I.E/Produtor': str  # Corrigido para corresponder ao nome exato da coluna no cadastro
}

# Carrega o arquivo Excel com as colunas corretas como string
df = pd.read_excel(file_name, dtype=col_types)

# Configurações adicionais para evitar notação científica
pd.set_option('display.float_format', '{:,.2f}'.format)
pd.options.display.float_format = '{:,.0f}'.format
pd.set_option('display.max_colwidth', None)

def adicionar_resultados(df, resultados):
    """Adiciona os resultados à tabela existente e ordena alfabeticamente"""
    df = pd.concat([df, resultados], ignore_index=True)
    df = df.sort_values(by='Produtor/Empresa')  # Ordena por Produtor/Empresa (alfabética)
    df.to_excel(file_name, index=False)
    print("Resultados adicionados à tabela existente com sucesso!")

def destacar_linhas(df, resultados):
    """Destaca as linhas encontradas na tabela com cor vermelha"""
    # Abre o arquivo Excel existente
    wb = load_workbook(file_name)
    ws = wb.active
    
    # Define o formato para destacar (vermelho)
    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Para cada linha encontrada, aplica o destaque
    for idx in resultados.index:  # Usamos o índice original do DataFrame `df`
        for col_idx in range(1, len(df.columns) + 1):  # Colunas começam de 1 no Excel
            ws.cell(row=idx + 2, column=col_idx).fill = fill  # +2 porque o Excel começa da linha 2
    
    # Salva o arquivo com as alterações de destaque
    wb.save(file_name)
    print("Linhas destacadas em vermelho!")

def abrir_excel(file_name):
    """Abre o arquivo Excel automaticamente após salvar ou finalizar a pesquisa"""
    if os.name == 'posix':  # Para sistemas Linux/Mac
        subprocess.call(['open', file_name])
    elif os.name == 'nt':  # Para sistemas Windows
        os.startfile(file_name)

def pesquisar_entregas():
    while True:
        criterio = input("Deseja pesquisar por (1) Produtor/Empresa, (2) CPF/CNPJ ou (3) Data? (digite 1, 2 ou 3): ")
        if criterio not in ['1', '2', '3']:
            print("Opção inválida. Tente novamente.")
            continue

        valor = input("Digite o valor de pesquisa: ")
        if criterio == '1':
            resultados = df[df['Produtor/Empresa'].str.contains(valor, case=False, na='')]
        elif criterio == '2':
            resultados = df[df['CPF/CNPJ'].str.contains(valor, na='')]
        elif criterio == '3':
            if '-' in valor:  # Caso tenha um "-"
                try:
                    data_inicio, data_fim = valor.split('-')
                    data_inicio = pd.to_datetime(data_inicio.strip(), errors='coerce')
                    data_fim = pd.to_datetime(data_fim.strip(), errors='coerce')

                    if pd.isna(data_inicio) or pd.isna(data_fim):
                        print("Formato de data inválido. Tente novamente.")
                        continue

                    resultados = df[(pd.to_datetime(df['Data'], errors='coerce') >= data_inicio) & 
                                    (pd.to_datetime(df['Data'], errors='coerce') <= data_fim)]
                except Exception as e:
                    print(f"Erro ao processar o intervalo de datas: {e}")
                    continue
            else:
                data_pesquisa = pd.to_datetime(valor.strip(), errors='coerce')
                if pd.isna(data_pesquisa):
                    print("Formato de data inválido. Tente novamente.")
                    continue
                resultados = df[pd.to_datetime(df['Data'], errors='coerce') == data_pesquisa]

        # Exibe os resultados
        if not resultados.empty:
            print("\nResultados encontrados:")
            print(resultados)

            # Calcula o total de entregas para este produtor/período
            total_entregas = len(resultados)
            
            # Calcula o total de embalagens por categoria
            total_lavadas = resultados[['Embalagens Lavadas 1L', 'Embalagens Lavadas 5L', 
                                      'Embalagens Lavadas 10L', 'Embalagens Lavadas 20L',
                                      'Embalagens Lavadas ML', 'Embalagens Lavadas GR', 
                                      'Embalagens Lavadas KG']].sum().sum()
            
            total_nao_lavadas = resultados[['Embalagens Não Lavadas 1L', 'Embalagens Não Lavadas 5L', 
                                          'Embalagens Não Lavadas 10L', 'Embalagens Não Lavadas 20L',
                                          'Embalagens Não Lavadas ML', 'Embalagens Não Lavadas GR', 
                                          'Embalagens Não Lavadas KG']].sum().sum()
            
            total_improprias = resultados[['Embalagens Impróprias 1L', 'Embalagens Impróprias 5L', 
                                         'Embalagens Impróprias 10L', 'Embalagens Impróprias 20L',
                                         'Embalagens Impróprias ML', 'Embalagens Impróprias GR', 
                                         'Embalagens Impróprias KG']].sum().sum()

            total_geral = total_lavadas + total_nao_lavadas + total_improprias

            print("\nResumo das entregas:")
            print(f"Total de entregas realizadas: {total_entregas}")
            print(f"\nTotal de embalagens entregues: {total_geral}")
            print(f"- Embalagens Lavadas: {total_lavadas}")
            print(f"- Embalagens Não Lavadas: {total_nao_lavadas}")
            print(f"- Embalagens Impróprias: {total_improprias}")

            salvar_tabela = input("Deseja salvar esses resultados em uma nova tabela? (s/n): ")
            if salvar_tabela.lower() == 's':
                novo_arquivo = f"resultados_pesquisa_{valor}.xlsx"
                resultados.to_excel(novo_arquivo, index=False)
                print(f"Resultados salvos em '{novo_arquivo}'")
                abrir_excel(novo_arquivo)

            adicionar = input("Deseja adicionar esses resultados à tabela existente? (s/n): ")
            if adicionar.lower() == 's':
                adicionar_resultados(df, resultados)

            destacar = input("Deseja destacar essas linhas em vermelho na tabela original? (s/n): ")
            if destacar.lower() == 's':
                destacar_linhas(df, resultados)

            editar = input("Deseja editar algum registro? (s/n): ")
            if editar.lower() == 's':
                index = int(input("Digite o número do registro que deseja editar (começando de 0): "))
                if 0 <= index < len(resultados):
                    for coluna in df.columns:
                        novo_valor = input(f"Novo valor para {coluna} (deixe em branco para manter '{resultados.iloc[index][coluna]}'): ")
                        if novo_valor:
                            df.at[resultados.index[index], coluna] = novo_valor
                    print("Registro atualizado com sucesso!")
                else:
                    print("Índice inválido.")
        else:
            print("Nenhum resultado encontrado.")

        continuar = input("Deseja realizar outra pesquisa? (s/n): ")
        if continuar.lower() != 's':
            break

    # Salva as alterações no arquivo Excel original
    df.to_excel(file_name, index=False)
    print("Alterações salvas com sucesso!")

    # Abre o arquivo Excel após salvar
    abrir_excel(file_name)

# Chama a função de pesquisa
pesquisar_entregas()
